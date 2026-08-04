from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from app.database import get_db
from app.core.permissions import get_current_user
from app.models.user import User
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from fastapi.responses import StreamingResponse
import pytz

IST = pytz.timezone('Asia/Kolkata')

router = APIRouter(prefix="/api/attendance/faculty", tags=["attendance_matrix"])

@router.get("/matrix")
async def get_attendance_matrix(
    date: date = Query(...),
    group_by: str = Query("all"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.session import AttendanceSession
    from app.models.user import User
    from app.models.attendance import AttendanceRecord
    
    # 1. Fetch all sessions for this faculty on this date
    stmt_sessions = select(AttendanceSession).where(
        AttendanceSession.faculty_id == current_user.id
    ).order_by(AttendanceSession.start_time)
    
    res = await db.execute(stmt_sessions)
    all_sessions = res.scalars().all()
    
    day_sessions = [s for s in all_sessions if s.start_time and s.start_time.date() == date]
    
    if not day_sessions:
        return {"sessions": [], "students": [], "grouped_students": {}}
        
    from app.models.rbac import Role
    stmt_students = select(User).join(Role, User.role_id == Role.id).where(Role.name == 'student').options(selectinload(User.profile))
    res_students = await db.execute(stmt_students)
    students = res_students.scalars().all()
    
    session_ids = [s.id for s in day_sessions]
    
    stmt_att = select(AttendanceRecord).where(
        AttendanceRecord.session_id.in_(session_ids)
    )
    res_att = await db.execute(stmt_att)
    attendances = res_att.scalars().all()
    
    # map: student_id -> session_id -> bool
    att_map = {}
    for a in attendances:
        if a.student_id not in att_map:
            att_map[a.student_id] = {}
        att_map[a.student_id][a.session_id] = (a.status == "present")
        
    student_list = []
    for s in students:
        s_data = {
            "id": s.id,
            "campus_id": s.campus_id,
            "name": (s.profile.name if s.profile else 'N/A'),
            "email": s.email,
            "phone": (s.profile.phone if s.profile else 'N/A'),
            "course": (s.profile.course if s.profile else 'N/A'),
            "specialisation": (s.profile.specialisation if s.profile else 'N/A'),
            "semester": (s.profile.semester if s.profile else 'N/A'),
            "attendance": {sid: att_map.get(s.id, {}).get(sid, False) for sid in session_ids}
        }
        student_list.append(s_data)
        
    # Group students based on group_by flag
    grouped_students = {}
    if group_by == "course":
        for s in student_list:
            # e.g., "BCA - Data Science (Semester I)"
            spec = s["specialisation"] if s["specialisation"] else "General"
            group_key = f"{s['course']} - {spec} (Sem {s['semester']})"
            if group_key not in grouped_students:
                grouped_students[group_key] = []
            grouped_students[group_key].append(s)
    else:
        # Default: group all students together
        grouped_students = {"All Students": student_list}
        
    sessions_data = []
    for s in day_sessions:
        if s.start_time:
            st = s.start_time
            if st.tzinfo is None:
                st = st.replace(tzinfo=pytz.UTC)
            ist_start = st.astimezone(IST)
        else:
            ist_start = None
            
        if s.end_time:
            et = s.end_time
            if et.tzinfo is None:
                et = et.replace(tzinfo=pytz.UTC)
            ist_end = et.astimezone(IST)
        else:
            ist_end = None

        time_str = "Unknown"
        if getattr(s, "time_slot", None):
            slot_map = {
                "09:30-11:00": "09:30 AM - 11:00 AM",
                "11:00-12:30": "11:00 AM - 12:30 PM",
                "01:30-02:30": "01:30 PM - 02:30 PM",
                "01:30-03:30": "01:30 PM - 03:30 PM",
                "02:00-03:30": "02:00 PM - 03:30 PM",
                "02:30-03:30": "02:30 PM - 03:30 PM"
            }
            time_str = slot_map.get(s.time_slot, s.time_slot)
        else:
            if ist_start and ist_end:
                time_str = f"{ist_start.strftime('%I:%M %p')} - {ist_end.strftime('%I:%M %p')}"
            elif ist_start:
                time_str = ist_start.strftime('%I:%M %p')
            
        sessions_data.append({
            "id": s.id,
            "subject_id": s.subject_id,
            "start_time": time_str
        })
    
    unique_dates = []
    for s in all_sessions:
        if s.start_time:
            d = s.start_time.date()
            if d not in unique_dates:
                unique_dates.append(d)
            
    try:
        day_number = unique_dates.index(date) + 1
    except ValueError:
        day_number = "?"

    return {
        "day_number": day_number,
        "sessions": sessions_data,
        "grouped_students": grouped_students
    }


def _make_header_style(bold=True, size=11, color="FFFFFF", bg_color=None, center=True):
    """Helper to create cell style."""
    font = Font(bold=bold, size=size, color=color)
    alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    fill = PatternFill("solid", fgColor=bg_color) if bg_color else None
    return font, alignment, fill


def _apply_border(cell):
    thin = Side(style="thin", color="4A5568")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


@router.get("/export/excel")
async def export_excel(
    date: date = Query(...),
    group_by: str = Query("all"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    matrix_data = await get_attendance_matrix(date=date, db=db, current_user=current_user, group_by=group_by)
    
    wb = openpyxl.Workbook()
    
    if not matrix_data["sessions"]:
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No sessions found for this date."])
    else:
        sessions = matrix_data["sessions"]
        n_sessions = len(sessions)
        # Fixed columns: S.No, Campus ID, Name, Email, Phone → 5 columns
        FIXED_COLS = 5
        total_cols = FIXED_COLS + n_sessions
        
        first = True
        for group_name, students in matrix_data["grouped_students"].items():
            safe_title = group_name.replace("/", "-").replace("[", "").replace("]", "").replace(":", "")[:31]
            if first:
                ws = wb.active
                ws.title = safe_title
                first = False
            else:
                ws = wb.create_sheet(title=safe_title)

            # ── Row 1: Group title (e.g. BCA(Data Science)) ──────────────────
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            c = ws.cell(row=1, column=1, value=group_name)
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor="1E3A5F")
            ws.row_dimensions[1].height = 28

            # ── Row 2: Total count ────────────────────────────────────────────
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
            c2 = ws.cell(row=2, column=1, value=f"Total Students Registered: {len(students)}")
            c2.font = Font(italic=True, size=11, color="94A3B8")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.fill = PatternFill("solid", fgColor="0F172A")
            ws.row_dimensions[2].height = 20

            # ── Row 3: "Day 1" super-header merged over session columns ───────
            # Fixed columns (S.No .. Phone) get no Day label; merge them vertically with row 4
            fixed_headers = ["S.No", "Campus ID", "Name", "Email", "Phone Number"]
            for col_idx, header in enumerate(fixed_headers, 1):
                # Write to the top-left cell of the merge (row 3)
                c = ws.cell(row=3, column=col_idx, value=header)
                c.font = Font(bold=True, size=10, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor="334155")
                _apply_border(c)
                # Apply border and fill to the bottom cell of the merge too for visual consistency
                c_bottom = ws.cell(row=4, column=col_idx)
                c_bottom.fill = PatternFill("solid", fgColor="334155")
                _apply_border(c_bottom)
                ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
            
            # Day header spanning the session columns
            ws.merge_cells(
                start_row=3,
                start_column=FIXED_COLS + 1,
                end_row=3,
                end_column=total_cols
            )
            day_num = matrix_data.get("day_number", "?")
            day_cell = ws.cell(row=3, column=FIXED_COLS + 1, value=f"Day {day_num}")
            day_cell.font = Font(bold=True, size=12, color="FFFFFF")
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.fill = PatternFill("solid", fgColor="1D4ED8")
            ws.row_dimensions[3].height = 22

            # ── Row 4: Column sub-headers (Sessions) ──────────────────────────


            for s_idx, s in enumerate(sessions, 1):
                col = FIXED_COLS + s_idx
                c = ws.cell(row=4, column=col, value=f"Session {s_idx}\n{s['start_time']}")
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.fill = PatternFill("solid", fgColor="334155")
                _apply_border(c)
                ws.row_dimensions[4].height = 40

            # ── Rows 5+: Data ─────────────────────────────────────────────────
            for i, user_obj in enumerate(students, 1):
                row_num = 4 + i
                row_data = [
                    i,
                    user_obj["campus_id"],
                    user_obj["name"],
                    user_obj["email"],
                    user_obj["phone"] or "N/A"
                ]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(size=10)
                    _apply_border(c)

                for s_idx, s in enumerate(sessions, 1):
                    col = FIXED_COLS + s_idx
                    present = user_obj["attendance"].get(s["id"], False)
                    c = ws.cell(row=row_num, column=col, value="✅" if present else "❌")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(size=12)
                    _apply_border(c)

            # ── Column widths ─────────────────────────────────────────────────
            col_widths = [8, 14, 22, 28, 14] + [22] * n_sessions
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Attendance_{date.isoformat()}.xlsx"}
    )

@router.get("/export/feedback/excel/{session_id}")
async def export_feedback_excel(
    session_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.session import AttendanceSession
    from app.models.user import User
    from app.models.attendance import AttendanceRecord
    from app.models.student_session_detail import StudentSessionDetail
    from datetime import timezone, timedelta
    
    IST = timezone(timedelta(hours=5, minutes=30))
    
    # Verify session belongs to faculty
    session_res = await db.execute(select(AttendanceSession).where(
        AttendanceSession.id == session_id,
        AttendanceSession.faculty_id == current_user.id
    ))
    session_obj = session_res.scalars().first()
    
    if not session_obj:
        raise HTTPException(status_code=404, detail="Session not found")
        
    # Query all students who attended and their feedback
    stmt = (
        select(Student, StudentSessionDetail)
        .join(AttendanceRecord, User.id == AttendanceRecord.student_id)
        .join(StudentSessionDetail, AttendanceRecord.id == StudentSessionDetail.attendance_id)
        .where(AttendanceRecord.session_id == session_id)
        .order_by(StudentSessionDetail.submitted_time)
    )
    
    res = await db.execute(stmt)
    records = res.all()
    
    # Group by course+specialisation
    grouped_feedbacks = {}
    for user_obj, feedback in records:
        key = f"{(user_obj.profile.course if user_obj.profile else 'N/A') or 'Unknown'}({(user_obj.profile.specialisation if user_obj.profile else 'N/A') or 'None'})"
        if key not in grouped_feedbacks:
            grouped_feedbacks[key] = []
        grouped_feedbacks[key].append((user_obj, feedback))
        
    wb = openpyxl.Workbook()
    
    if not grouped_feedbacks:
        ws = wb.active
        ws.title = f"Feedback Session {session_id}"
        ws.append(["No feedback found."])
    else:
        first = True
        for group_name, items in grouped_feedbacks.items():
            safe_title = group_name.replace("/", "-").replace("[", "").replace("]", "").replace(":", "")[:31]
            if first:
                ws = wb.active
                ws.title = safe_title
                first = False
            else:
                ws = wb.create_sheet(title=safe_title)
                
            # Headers
            headers = [
                "S.No", "Campus ID", "Student Name", 
                "Interactive Rating", "Relevant Rating", 
                "Learned Today", "Key Takeaway", 
                "Overall Satisfaction", "Issue Note", "Submitted At"
            ]
            
            # Styling
            thin_border = Border(
                left=Side(style='thin', color='475569'),
                right=Side(style='thin', color='475569'),
                top=Side(style='thin', color='475569'),
                bottom=Side(style='thin', color='475569')
            )
            
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=1, column=col_idx, value=header)
                c.font = Font(bold=True, size=11, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor="334155")
                c.border = thin_border
                
            ws.row_dimensions[1].height = 25
            
            for i, (user_obj, feedback) in enumerate(items, 1):
                row_num = i + 1
                sub_time = "N/A"
                if feedback.submitted_time:
                    dt = feedback.submitted_time
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    sub_time = dt.astimezone(IST).strftime('%I:%M %p')
                    
                row_data = [
                    i,
                    user_obj.campus_id,
                    (user_obj.profile.name if user_obj.profile else 'N/A'),
                    f"{feedback.interactive_rating}/5" if feedback.interactive_rating else "N/A",
                    f"{feedback.relevant_rating}/5" if feedback.relevant_rating else "N/A",
                    feedback.learned_today or "N/A",
                    feedback.key_takeaway or "N/A",
                    f"{feedback.overall_satisfaction}/5" if feedback.overall_satisfaction else "N/A",
                    feedback.issue_note or "N/A",
                    sub_time
                ]
                
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = thin_border
                    
            # Column widths
            col_widths = [6, 12, 22, 16, 16, 30, 30, 18, 30, 14]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Use time_slot for filename if available
    time_str = session_obj.time_slot.replace(":", "-") if getattr(session_obj, "time_slot", None) else str(session_id)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Feedback_Session_{time_str}.xlsx"}
    )

@router.get("/export/feedback/excel/date/{date_str}")
async def export_feedback_excel_by_date(
    date_str: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.session import AttendanceSession
    from app.models.user import User
    from app.models.attendance import AttendanceRecord
    from app.models.student_session_detail import StudentSessionDetail
    from datetime import datetime, timezone, timedelta
    
    IST = timezone(timedelta(hours=5, minutes=30))
    
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format, expected YYYY-MM-DD")
        
    start_dt = datetime.combine(date_obj, datetime.min.time()).replace(tzinfo=IST)
    end_dt = datetime.combine(date_obj, datetime.max.time()).replace(tzinfo=IST)
    
    start_utc = start_dt.astimezone(timezone.utc)
    end_utc = end_dt.astimezone(timezone.utc)
    
    # Verify sessions belong to faculty on this date
    session_res = await db.execute(select(AttendanceSession).where(
        AttendanceSession.faculty_id == current_user.id,
        AttendanceSession.start_time >= start_utc,
        AttendanceSession.start_time <= end_utc
    ))
    sessions = session_res.scalars().all()
    
    if not sessions:
        raise HTTPException(status_code=404, detail="No sessions found for this date")
        
    session_ids = [s.id for s in sessions]
    session_map = {s.id: s for s in sessions}
    
    # Query all students who attended and their feedback
    stmt = (
        select(Student, StudentSessionDetail, AttendanceRecord.session_id)
        .join(AttendanceRecord, User.id == AttendanceRecord.student_id)
        .join(StudentSessionDetail, AttendanceRecord.id == StudentSessionDetail.attendance_id)
        .where(AttendanceRecord.session_id.in_(session_ids))
        .order_by(StudentSessionDetail.submitted_time)
    )
    
    res = await db.execute(stmt)
    records = res.all()
    
    # Group by session_id
    grouped_feedbacks = {}
    for user_obj, feedback, sess_id in records:
        if sess_id not in grouped_feedbacks:
            grouped_feedbacks[sess_id] = []
        grouped_feedbacks[sess_id].append((user_obj, feedback))
        
    wb = openpyxl.Workbook()
    
    if not grouped_feedbacks:
        ws = wb.active
        ws.title = f"Feedbacks {date_str}"
        ws.append(["No feedback found for any session on this date."])
    else:
        first = True
        for sess_id, items in grouped_feedbacks.items():
            sess = session_map.get(sess_id)
            group_name = f"{sess.subject_id} - {sess.time_slot}" if sess and sess.subject_id else f"Session {sess_id}"
            safe_title = group_name.replace("/", "-").replace("[", "").replace("]", "").replace(":", "").replace("*", "").replace("?", "")[:31]
            if first:
                ws = wb.active
                ws.title = safe_title
                first = False
            else:
                ws = wb.create_sheet(title=safe_title)
                
            # Headers
            headers = [
                "S.No", "Campus ID", "Student Name", 
                "Interactive Rating", "Relevant Rating", 
                "Learned Today", "Key Takeaway", 
                "Overall Satisfaction", "Issue Note", "Submitted At"
            ]
            
            # Styling
            thin_border = Border(
                left=Side(style='thin', color='475569'),
                right=Side(style='thin', color='475569'),
                top=Side(style='thin', color='475569'),
                bottom=Side(style='thin', color='475569')
            )
            
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=1, column=col_idx, value=header)
                c.font = Font(bold=True, size=11, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor="334155")
                c.border = thin_border
                
            ws.row_dimensions[1].height = 25
            
            for i, (user_obj, feedback) in enumerate(items, 1):
                row_num = i + 1
                sub_time = "N/A"
                if feedback.submitted_time:
                    dt = feedback.submitted_time
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    sub_time = dt.astimezone(IST).strftime('%I:%M %p')
                    
                row_data = [
                    i,
                    user_obj.campus_id,
                    (user_obj.profile.name if user_obj.profile else 'N/A'),
                    f"{feedback.interactive_rating}/5" if feedback.interactive_rating else "N/A",
                    f"{feedback.relevant_rating}/5" if feedback.relevant_rating else "N/A",
                    feedback.learned_today or "N/A",
                    feedback.key_takeaway or "N/A",
                    f"{feedback.overall_satisfaction}/5" if feedback.overall_satisfaction else "N/A",
                    feedback.issue_note or "N/A",
                    sub_time
                ]
                
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = thin_border
                    
            # Column widths
            col_widths = [6, 12, 22, 16, 16, 30, 30, 18, 30, 14]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Feedbacks_{date_str}.xlsx"}
    )


async def get_consolidated_attendance_matrix(
    group_by: str = "course",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.session import AttendanceSession
    from app.models.user import User
    from app.models.attendance import AttendanceRecord
    
    # 1. Fetch all sessions for this faculty
    stmt_sessions = select(AttendanceSession).where(
        AttendanceSession.faculty_id == current_user.id
    ).order_by(AttendanceSession.start_time)
    
    res = await db.execute(stmt_sessions)
    all_sessions = res.scalars().all()
    
    if not all_sessions:
        return {"dates": [], "grouped_students": {}}
        
    from app.models.rbac import Role
    stmt_students = select(User).join(Role, User.role_id == Role.id).where(Role.name == 'student').options(selectinload(User.profile))
    res_students = await db.execute(stmt_students)
    students = res_students.scalars().all()
    
    session_ids = [s.id for s in all_sessions]
    
    stmt_att = select(AttendanceRecord).where(
        AttendanceRecord.session_id.in_(session_ids)
    )
    res_att = await db.execute(stmt_att)
    attendances = res_att.scalars().all()
    
    # map: student_id -> date -> bool (attended at least one session that day)
    att_map = {}
    
    # Pre-map session_id to date
    session_date_map = {}
    for s in all_sessions:
        session_date_map[s.id] = s.start_time.date()
        
    # Get unique dates in chronological order
    unique_dates = sorted(list(set(session_date_map.values())))

    for a in attendances:
        if a.student_id not in att_map:
            att_map[a.student_id] = {}
        s_date = session_date_map.get(a.session_id)
        if s_date:
            if a.status == "present":
                att_map[a.student_id][s_date] = True
        
    student_list = []
    for s in students:
        s_data = {
            "id": s.id,
            "campus_id": s.campus_id,
            "name": (s.profile.name if s.profile else 'N/A'),
            "email": s.email,
            "phone": (s.profile.phone if s.profile else 'N/A'),
            "course": (s.profile.course if s.profile else 'N/A'),
            "specialisation": (s.profile.specialisation if s.profile else 'N/A'),
            "semester": (s.profile.semester if s.profile else 'N/A'),
            "attendance": {d: att_map.get(s.id, {}).get(d, False) for d in unique_dates}
        }
        student_list.append(s_data)
        
    # Group students based on group_by flag
    grouped_students = {}
    if group_by == "course" or group_by == "specialisation":
        for s in student_list:
            spec = s["specialisation"] if s["specialisation"] else "General"
            group_key = f"{s['course']} - {spec} (Sem {s['semester']})"
            if group_key not in grouped_students:
                grouped_students[group_key] = []
            grouped_students[group_key].append(s)
    else:
        grouped_students = {"All Students": student_list}

    return {
        "dates": unique_dates,
        "grouped_students": grouped_students
    }

@router.get("/export/consolidated/excel")
async def export_consolidated_excel(
    group_by: str = Query("course"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    matrix_data = await get_consolidated_attendance_matrix(group_by=group_by, db=db, current_user=current_user)
    
    wb = openpyxl.Workbook()
    
    if not matrix_data["dates"]:
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No attendance data found."])
    else:
        dates = matrix_data["dates"]
        n_dates = len(dates)
        FIXED_COLS = 5
        total_cols = FIXED_COLS + n_dates
        
        first = True
        for group_name, students in matrix_data["grouped_students"].items():
            safe_title = group_name.replace("/", "-").replace("[", "").replace("]", "").replace(":", "")[:31]
            if first:
                ws = wb.active
                ws.title = safe_title
                first = False
            else:
                ws = wb.create_sheet(title=safe_title)

            # ── Row 1: Group title ──────────────────
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            c = ws.cell(row=1, column=1, value=f"{group_name} - Consolidated Daily Attendance")
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor="1E3A5F")
            ws.row_dimensions[1].height = 28

            # ── Row 2: Total count ────────────────────────────────────────────
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
            c2 = ws.cell(row=2, column=1, value=f"Total Students Registered: {len(students)}")
            c2.font = Font(italic=True, size=11, color="94A3B8")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.fill = PatternFill("solid", fgColor="0F172A")
            ws.row_dimensions[2].height = 20

            # ── Row 3: Headers ────────────────────────────────────────────
            fixed_headers = ["S.No", "Campus ID", "Name", "Email", "Phone Number"]
            for col_idx, header in enumerate(fixed_headers, 1):
                c = ws.cell(row=3, column=col_idx, value=header)
                c.font = Font(bold=True, size=10, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor="334155")
                _apply_border(c)
            
            for d_idx, d in enumerate(dates, 1):
                col = FIXED_COLS + d_idx
                c = ws.cell(row=3, column=col, value=f"Day {d_idx}\n{d.strftime('%b %d')}")
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.fill = PatternFill("solid", fgColor="1D4ED8")
                _apply_border(c)
            
            ws.row_dimensions[3].height = 30

            # ── Rows 4+: Data ─────────────────────────────────────────────────
            for i, user_obj in enumerate(students, 1):
                row_num = 3 + i
                row_data = [
                    i,
                    user_obj["campus_id"],
                    user_obj["name"],
                    user_obj["email"],
                    user_obj["phone"] or "N/A"
                ]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(size=10)
                    _apply_border(c)

                for d_idx, d in enumerate(dates, 1):
                    col = FIXED_COLS + d_idx
                    present = user_obj["attendance"].get(d, False)
                    c = ws.cell(row=row_num, column=col, value="✅" if present else "❌")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(size=12)
                    _apply_border(c)

            # ── Column widths ─────────────────────────────────────────────────
            col_widths = [8, 14, 22, 28, 14] + [12] * n_dates
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Consolidated_Attendance.xlsx"}
    )
